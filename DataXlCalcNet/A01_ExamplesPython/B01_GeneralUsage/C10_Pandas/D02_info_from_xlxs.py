
from openpyxl import load_workbook
import warnings
import sys


def get_info_from_xlsx(ActiveFolder, FileName):
    warnings.simplefilter(action='ignore', category=UserWarning) # suppress a useless openpyxl warning
    wb = load_workbook(filename=ActiveFolder  + '\\' + FileName) # tables cannot be accessed in read-only workbooks

    # List General Info
    print()
    print('<H1 Title="General Info">')
    print('Workbook name: ' + FileName)
    print('Path: ' + ActiveFolder)
    print('Master Workbook: Yes')
    print('Opened: Yes')
    print('Standard Action: Expand Array Formulas')
    print('Standard Action: Insert Stats Sheet')
    print('</H1>')

    # List Expandable Array Formulas
    print()
    print('<H1 Title="Expandable Array Formulas">')
    print('Array Formula: Output1')
    print('Array Formula: Output2')
    print('Array Formula: Output3')
    print('</H1>')

    # List Worksheets and Tables
    print()
    print('<H1 Title="Worksheets and Tables">')
    wb_sheetnames = wb.sheetnames
    for SheetName in wb_sheetnames:
        print('Worksheet:', SheetName)
        ws = wb[SheetName]
        for table in ws.tables:
            print('    Table:', table)
    print('</H1>')

    # List Named Ranges
    print()
    print('<H1 Title="Named Ranges">')
    defined_names = wb.defined_names.definedName
    for dfname in defined_names:
        print('Named Range:', dfname.name)
    print('</H1>')

    # Close the workbook after reading
    wb.close()



try:
    for arg in sys.argv:
        print(arg)
    print()
    print(sys.path)
    print()
    get_info_from_xlsx(sys.argv[1], sys.argv[2])


except Exception:
    import traceback
    print(traceback.format_exc())

