
#from xlcalcnet import gui
from openpyxl import load_workbook
import warnings, csv, os
import sqlite3
import datetime

def get_info_from_xlsx(FileName, ActiveFolder):
    warnings.simplefilter(action='ignore', category=UserWarning)
    # tables cannot be accessed in read-only workbooks
    wb = load_workbook(filename=ActiveFolder  + '\\' + FileName + '.xlsx', data_only=True) 
    print('Workbook name: ' + FileName)
    print('Path: ' + ActiveFolder)
    fname = ActiveFolder  + '\\' + FileName + '.db'
    if os.path.exists(fname): os.remove(fname)

    sqlite3.register_adapter(datetime.date, adapt_date_iso)
    sqlite3.register_adapter(datetime.datetime, adapt_datetime_iso)

    # Connect to the SQLite database
    conn = sqlite3.connect(fname)
    cursor = conn.cursor()
    print()

    # Create the table definitions from the worksheet 'SQLDef'
    ws_name = 'SQLDef'
    datatablesheetlist = []
    print(f"worksheet name: {ws_name}")
    ws = wb[ws_name]
    print(f"tables in worksheet: {len(ws.tables)}")
    print()
    for tbl in ws.tables.values():
        print(f"table name: {tbl.name}")
        datatablesheetlist.append(tbl.name[:-len(ws_name)])
        data = ws[tbl.ref]
        rows_list = []
        first = True
        for row in data:
            if not first:
                cols = []
                for col in row: cols.append(col.value)
                rows_list.append(cols)
            else:
                first = False
        SQLDefStr = '\n'.join(flatten(rows_list))
        print(str(SQLDefStr))
        cursor.execute(SQLDefStr)
        conn.commit()
    print()

    # Write the data into the tables
    for ws_name in datatablesheetlist:
        print(ws_name)
#        if (ws_name == 'Hemodynamic'):
#        if (ws_name == 'Clinical'):
#        if (ws_name == 'Laboratory'):
#        if (ws_name == 'Patient'):
        if (True):
            ws = wb[ws_name]
            tbl = ws.tables[ws_name+'_Data']
            print(f"table name: {tbl.name}")
            data = ws[tbl.ref]
            first = True
            s = '?, ' * (len(tbl.tableColumns)-1)
            statement = 'INSERT INTO ' + ws_name + ' VALUES (' + s + '?)'
            print(statement)
            for row in data:
                if not first:
                    cols = []
                    for col in row: 
                        temp = col.value
                        if isinstance(temp, datetime.datetime):
                            adapt_datetime_iso(temp)
                        cols.append(temp)
                    print(cols)
                    cursor.execute(statement, cols)
                else: first = False
            conn.commit()


    # Close the database connection
    conn.close()

    # Close the workbook after reading
    wb.close()





def get_info_from_xlsx_old(FileName, ActiveFolder):
    warnings.simplefilter(action='ignore', category=UserWarning)
    # tables cannot be accessed in read-only workbooks
    wb = load_workbook(filename=ActiveFolder  + '\\' + FileName + '.xlsx', data_only=True) 

    # List General Info
    print()
    print('<H1 Title="General Info">')
    print('Workbook name: ' + FileName)
    print('Path: ' + ActiveFolder)
    print('</H1>')

    # List Worksheets and Tables
    print()
    print('<H1 Title="Worksheets and Tables">')

    fname = ActiveFolder  + '\\' + FileName + '.db'
    if os.path.exists(fname):
        os.remove(fname)

    # Connect to the SQLite database
    conn = sqlite3.connect(fname)
    cursor = conn.cursor()


    # Initialize the dictionary of tables
    tables_dict = {}

    # Go through each worksheet in the workbook
    for ws_name in wb.sheetnames:
        print("")
        print(f"worksheet name: {ws_name}")
        ws = wb[ws_name]
        print(f"tables in worksheet: {len(ws.tables)}")

        # Get each table in the worksheet
        for tbl in ws.tables.values():
            print(f"table name: {tbl.name}")
            # First, add some info about the table to the dictionary
            tables_dict[tbl.name] = {
                'table_name': tbl.name,
                'worksheet': ws_name,
                'num_cols': len(tbl.tableColumns),
                'table_range': tbl.ref}

            # Grab the 'data' from the table
            data = ws[tbl.ref]
            # First get a list of all rows, including the first header row
            rows_list = []
            first = True
            for row in data:
                if not first:
                    # Get a list of all columns in each row
                    cols = []
                    for col in row:
                        cols.append(col.value)
                    rows_list.append(cols)
                else:
                    first = False
            #print(rows_list)
            if tbl.name.endswith('SQLDef'):
                SQLDef = '\n'.join(flatten(rows_list))
                print(str(SQLDef))
                # Create a table for the data
                cursor.execute(SQLDef)

            else: print(rows_list)

        # Commit the changes and close the database connection
        conn.commit()
    conn.close()
            

    print('</H1>')

    # Close the workbook after reading
    wb.close()






def flatten(xss):
    return [x for xs in xss for x in xs]
def convert_datetime(val):
    """Convert ISO 8601 datetime to datetime.datetime object."""
    return datetime.datetime.fromisoformat(val.decode())
def adapt_date_iso(val):
    """Adapt datetime.date to ISO 8601 date."""
    return val.isoformat()
def adapt_datetime_iso(val):
    """Adapt datetime.datetime to timezone-naive ISO 8601 date."""
    return val.replace(tzinfo=None).isoformat()

def demo_cardiac_info():
    print('Hello from demo_cardiac!')
    FileName = 'Cardiac'
    ActiveFolder = r'C:\Users\DUHad\Documents\DataXlCalcNet\DataExamples\MainExamples\Workbooks'
    get_info_from_xlsx(FileName, ActiveFolder)

def demo_cardiac_info_old():
    print('Hello from demo_cardiac!')
    FileName = 'Cardiac'
    ActiveFolder = r'C:\Users\DUHad\Documents\DataXlCalcNet\DataExamples\MainExamples\Workbooks'
    get_info_from_xlsx_old(FileName, ActiveFolder)




try:
    demo_cardiac_info()


except Exception:
    import traceback
    print(traceback.format_exc())

